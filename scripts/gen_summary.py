# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
import json, csv, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUTS_ROOT = sys.argv[1]
BATCH_DIR = sys.argv[2]

HEADERS = [
    "序号", "视频ID", "标题/文案", "链接", "作者",
    "点赞数", "评论数", "分享数", "收藏数", "播放数",
    "时长(秒)", "发布日期", "转写状态", "原始文案全文", "本地文件路径"
]

rows = []
dirs = [d for d in os.listdir(OUTPUTS_ROOT)
        if "[douyin]" in d and os.path.isdir(os.path.join(OUTPUTS_ROOT, d))]

for d in sorted(dirs):
    dpath = os.path.join(OUTPUTS_ROOT, d)
    manifest_path = os.path.join(dpath, "_meta", "manifest.json")
    if not os.path.exists(manifest_path):
        continue

    with open(manifest_path, "r", encoding="utf-8") as f:
        m = json.load(f)

    # Find transcript - use listdir (glob fails on Windows with ? in path)
    transcript = ""
    try:
        for fname in os.listdir(dpath):
            if fname.endswith("原始逐字稿.txt"):
                tpath = os.path.join(dpath, fname)
                with open(tpath, "r", encoding="utf-8") as f:
                    transcript = f.read().strip()
                break
    except Exception:
        pass

    c = m.get("content", {})
    stats = c.get("stats", {})
    upload_date = c.get("upload_date", "")
    date_str = f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:8]}" if len(upload_date) == 8 else upload_date
    author = m.get("uploader", "")

    rows.append(dict(
        aweme_id=m.get("id", ""), desc=m.get("title", ""),
        share_url=m.get("webpage_url", ""), author=author,
        digg_count=stats.get("like_count", 0) or 0,
        comment_count=stats.get("comment_count", 0) or 0,
        share_count=stats.get("share_count", 0) or 0,
        collect_count=stats.get("favorite_count", 0) or 0,
        play_count=stats.get("view_count", 0) or 0,
        duration_sec=round(m.get("duration", 0), 1),
        date_str=date_str, transcript=transcript, output_dir=dpath,
    ))

rows.sort(key=lambda r: r["digg_count"], reverse=True)

qualifying_count = 0
qf = os.path.join(BATCH_DIR, "qualifying_videos.json")
if os.path.exists(qf):
    with open(qf, "r", encoding="utf-8") as f:
        qualifying_count = len(json.load(f))
print(f"Found {len(rows)} processed videos (threshold: {qualifying_count} qualifying)")

# Build rows
excel_rows = []
for i, r in enumerate(rows, 1):
    excel_rows.append([
        i, r["aweme_id"], r["desc"], r["share_url"], r["author"],
        r["digg_count"], r["comment_count"], r["share_count"],
        r["collect_count"], r["play_count"], r["duration_sec"],
        r["date_str"],
        "成功" if r["transcript"] else "缺文案",
        r["transcript"], r["output_dir"],
    ])

# Excel
wb = Workbook()
ws = wb.active
ws.title = "抖音文案汇总"

hfont = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
bd = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
for ci, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bd

cfont = Font(name="微软雅黑", size=10)
for ri, rd in enumerate(excel_rows, 2):
    for ci, val in enumerate(rd, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = cfont
        c.alignment = Alignment(vertical="center", wrap_text=(ci in [14, 15]))
        c.border = bd

for i, w in enumerate([6, 20, 35, 35, 14, 10, 10, 10, 10, 10, 10, 12, 10, 50, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

xlsx_path = os.path.join(BATCH_DIR, "汇总表.xlsx")
wb.save(xlsx_path)

csv_path = os.path.join(BATCH_DIR, "汇总表.csv")
with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(HEADERS)
    writer.writerows(excel_rows)

print(f"Excel: {xlsx_path}")
print(f"CSV: {csv_path}")
print(f"Rows: {len(excel_rows)}")
